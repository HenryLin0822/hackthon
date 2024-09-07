from openpyxl import load_workbook

def load(filename):
    # Load the workbook
    workbook = load_workbook(filename=filename)

    # Select the active sheet
    sheet = workbook.active

    store = []
    # Iterate through rows and access cell values
    for row in sheet.iter_rows(values_only=True):
        #print(row)
        store.append(row)
    return store

store = load("../大安森林公園.xlsx")
print(store[0][0]) #Prompt
print(store[0][1]) #Answer
print(store[1][0]) #請推薦捷運大安森林公園站附近美食
print(len(store)) #140